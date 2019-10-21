from openpyxl import load_workbook

path = r'D:\pythonT\fo.xlsx'
workbook = load_workbook(path)
sheet = workbook.get_sheet_by_name('Sheet2')
sheet.cell(1, 1, '哈哈')
workbook.save(path)

