# 工作簿的英文是BOOK（WORKBOOK），工作表的英文是SHEET（WORKSHEET）。
# 一个工作簿就是一个独立的文件
# 一个工作簿里面可以有1个或者多个工作表
# 工作簿是工作表的集合
import xlrd

# 设置路径
path = 'E:/input.xlsx'
# 打开execl
workbook = xlrd.open_workbook(path)

# 输出Excel文件中所有sheet的名字
print(workbook.sheet_names())

# 根据sheet索引或者名称获取sheet内容
Data_sheet = workbook.sheets()[0]  # 通过索引获取
# Data_sheet = workbook.sheet_by_index(0)  # 通过索引获取
# Data_sheet = workbook.sheet_by_name(u'名称')  # 通过名称获取


print(Data_sheet.name)  # 获取sheet名称
rowNum = Data_sheet.nrows  # sheet行数
colNum = Data_sheet.ncols  # sheet列数

# 获取所有单元格的内容
list = []
for i in range(rowNum):
    rowlist = []
    for j in range(colNum):
        rowlist.append(Data_sheet.cell_value(i, j))
    list.append(rowlist)
# 输出所有单元格的内容
for i in range(rowNum):
    for j in range(colNum):
        print(list[i][j], '\t\t', end="")
    print()

# 获取整行和整列的值（列表）
rows = Data_sheet.row_values(0)  # 获取第一行内容
cols = Data_sheet.col_values(1)  # 获取第二列内容
# print (rows)
# print (cols)

# 获取单元格内容
cell_A1 = Data_sheet.cell(0, 0).value
cell_B1 = Data_sheet.row(0)[1].value  # 使用行索引
cell_C1 = Data_sheet.cell(0, 2).value
cell_D2 = Data_sheet.col(3)[1].value  # 使用列索引
print(cell_A1, cell_B1, cell_C1, cell_D2)

# 获取单元格内容的数据类型
# ctype:0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
print('cell(0,0)数据类型:', Data_sheet.cell(0, 0).ctype)
print('cell(1,0)数据类型:', Data_sheet.cell(1, 0).ctype)
print('cell(1,1)数据类型:', Data_sheet.cell(1, 1).ctype)
print('cell(1,2)数据类型:', Data_sheet.cell(1, 2).ctype)

# 获取单元格内容为日期的数据
date_value = xlrd.xldate_as_tuple(Data_sheet.cell_value(1,0),workbook.datemode)
print(type(date_value), date_value)
print('%d:%d:%d' % (date_value[0:3]))

import xlwt


def set_style(name, height, bold=False):
    style = xlwt.XFStyle()   # 初始化样式
    font = xlwt.Font()       # 为样式创建字体
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height

    style.font = font
    return style


def write_excel(path):
    # 创建工作簿
    workbook = xlwt.Workbook(encoding='utf-8')
    # 创建sheet
    data_sheet = workbook.add_sheet('demo')
    # 注意：如果对一个单元格重复操作，会引发error。所以在打开时加cell_overwrite_ok=True解决
    # data_sheet = workbook.add_sheet('demo', cell_overwrite_ok=True)
    row0 = [u'字段名称', u'大致时段', 'CRNTI', 'CELL-ID']
    row1 = [u'测试', '15:50:33-15:52:14', 22706, 4190202]
    # 生成第一行和第二行
    for i in range(len(row0)):
        data_sheet.write(0, i, row0[i], set_style('Times New Roman', 220, True))
        data_sheet.write(1, i, row1[i], set_style('Times New Roman', 220, True))

    # 保存文件
    # workbook.save('demo.xls')
    workbook.save(path)


# 如果excel文件已经存在可以直接添加单元格数据
from openpyxl import load_workbook

path = r'D:\pythonT\f1.xlsx'
workbook = load_workbook(path)
sheet = workbook.get_sheet_by_name('Sheet2')
sheet.cell(1, 1, '哈哈')
workbook.save(path)

# pandas更为简单
# 读取
import pandas as pd
df = pd.read_excel(r'data.xlsx',sheetname=0)
print(df.head())
# 写
from pandas import DataFrame
data={
'name':['张三','李四','王五'],
'age':[11,12,13],
'sex':['男','女','男']
}
df= DataFrame(data)
df.to_excel('new.xlsx')




if __name__ == '__main__':
    # 设置路径
    path = 'E:/demo.xls'
    write_excel(path)
    print(u'创建demo.xls文件成功')